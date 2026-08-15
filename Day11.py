'''from openpyxl import Workbook

# Nayi Excel workbook banao (poori file)
wb = Workbook()

# Active sheet lo (ek tab/page)
ws = wb.active
ws.title = "Mera Sheet"

# Cells mein data likho (A1, B1 = dabbe ke naam)
ws["A1"] = "Naam"
ws["B1"] = "Age"

ws["A2"] = "Vanshika"
ws["B2"] = 21

# File save karo
wb.save("pehla.xlsx")
print("✅ pehla.xlsx ban gayi! Folder mein khol kar dekho.")'''

# mini project - expense tracker 
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Nayi workbook
wb = Workbook()
ws = wb.active
ws.title = "Expenses"

# Headers (pehli row)
ws.append(["Date", "Item", "Category", "Amount"])

# Header ko bold + blue color karo
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Data rows (append = ek row jodo)
expenses = [
    ["2026-08-01", "Groceries", "Food", 2500],
    ["2026-08-02", "Petrol", "Transport", 1000],
    ["2026-08-03", "Movie", "Fun", 800],
    ["2026-08-04", "Bijli Bill", "Bills", 1500],
    ["2026-08-05", "Restaurant", "Food", 1200],
]
for row in expenses:
    ws.append(row)

# TOTAL ka formula (Excel khud jod-tod karega)
ws.append([])
ws.append(["", "", "TOTAL", "=SUM(D2:D6)"])

# Column width sundar karo
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 10

# Save
wb.save("expenses.xlsx")
print("✅ expenses.xlsx ban gayi! Excel mein kholo.") 


# Reading Excel 
from openpyxl import load_workbook

# Excel file kholo
wb = load_workbook("expenses.xlsx")
ws = wb.active

print("\nExcel ka data:")
for row in ws.iter_rows(values_only=True):
    print(row)