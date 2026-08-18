import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

print("🚀 Auto Report Generator Started...")

# ---------- 1. DATA LOAD & CLEAN ----------
df = pd.read_csv("messy_sales.csv")
print(f"Raw data rows: {len(df)}")

# Missing values (None/NaN) wali rows hatao
df = df.dropna()

# Duplicate rows hatao (jaise 2026-08-01 wala Laptop do baar tha)
df = df.drop_duplicates()
print(f"Clean data rows: {len(df)}")

# ---------- 2. DATA ANALYSIS ----------
# Region ke hisaab se Total Sales nikalo
summary = df.groupby("Region")["Sales"].sum().reset_index()
summary.columns = ["Region", "Total Sales"]

# Top Product nikalo
top_product = df.groupby("Product")["Sales"].sum().reset_index()
top_product = top_product.sort_values("Sales", ascending=False).head(1)

# ---------- 3. EXCEL EXPORT (Multiple Sheets) ----------
file_name = "final_report.xlsx"
with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Clean Data", index=False)
    summary.to_excel(writer, sheet_name="Region Summary", index=False)

# ---------- 4. EXCEL STYLING (Sundar banao) ----------
wb = load_workbook(file_name)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for cell in ws[1]: # Pehli row (headers)
        cell.font = header_font
        cell.fill = header_fill
    # Column width auto-adjust
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column].width = max_length + 2

wb.save(file_name)
print(f"✅ {file_name} successfully generated and styled!")

# ---------- 5. EMAIL SEND (Optional - Try/Except se safe) ----------
print("\n📧 Email bhejne ki koshish ho rahi hai...")

try:
    # Yahan apni details dalo (App password wala tarika)
    sender = input("Apna Gmail: ")
    app_password = input("App Password: ")
    receiver = input("Kisko bhejna hai: ")

    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = receiver
    msg["Subject"] = "Daily Sales Report 📊"
    
    # Email body mein Top product ka zikr
    body = f"Hello,\n\nAaj ki report attach hai.\n\n🏆 Top Product: {top_product.iloc[0]['Product']} (Sales: {top_product.iloc[0]['Sales']})\n\nRegards,\nAutomation Script"
    msg.attach(MIMEText(body, "plain"))

    # Attachment jodo
    with open(file_name, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={file_name}")
    msg.attach(part)

    # Server connect aur send
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(sender, app_password)
    server.sendmail(sender, receiver, msg.as_string())
    server.quit()
    print("✅ Email successfully bhej di gayi with attachment!")

except Exception as e:
    print(f"⚠️ Email skip ho gaya (Error: {e}). Par Excel report toh ban gayi hai!")

print("\n🎉 PROCESS COMPLETE!")