import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

REPORT_FILE = "internship_report.xlsx"

def generate_report(candidates, new_candidates):
    """Saari matched jobs ka Excel report banao"""
    if not candidates:
        print("📊 Report ke liye koi job nahi mili")
        return

    # New jobs ka ID set (quick lookup)
    new_ids = {c["id"] for c in new_candidates}

    # Data ready karo
    rows = []
    for c in candidates:
        rows.append({
            "New": "YES" if c["id"] in new_ids else "",
            "Title": c["title"],
            "Company": c["company"],
            "Mode": "WFH" if c["mode"] == "wfh" else "Office",
            "Location": c["location"],
            "Stipend": c["stipend"],
            "Match Score": c["total"],
            "Core Skills": ", ".join(c["matched_core"]),
            "Known Skills": ", ".join(c["matched_known"]),
            "Apply URL": c.get("url", ""),
        })

    df = pd.DataFrame(rows)
    df.to_excel(REPORT_FILE, index=False, sheet_name="Matched Jobs")

    # Style karo
    style_report()
    print(f"📊 Excel report ban gayi: {REPORT_FILE}")

def style_report():
    """Excel ko sundar banao (Day 11 wala jaadu!)"""
    wb = load_workbook(REPORT_FILE)
    ws = wb["Matched Jobs"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Header style karo
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths auto-adjust
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 45)

    wb.save(REPORT_FILE)